#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Nest Backend → مزامنة تقارير الوردية إلى شيت "السجل اليومي" في ملف Nest Strategy.
يعمل محلياً بالكامل: يقرأ من nest.db (SQLite) ويكتب صفوفاً جديدة في الشيت،
بدون أي اتصال خارجي أو صلاحيات Google — مجرد تعديل ملف Excel محلي.
يتجنب التكرار عبر عمود مخفي "ref_id" (Z) يربط كل صف بمعرّف تقرير الوردية.
"""
import sqlite3
import sys
from datetime import datetime, timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DB_PATH = "/home/claude/nest_backend/nest.db"
XLSX_PATH = "/home/claude/Nest_Strategy_Management.xlsx"
LOG_SHEET = "السجل اليومي"
REF_COL = 26  # column Z, hidden tracking column (لا يظهر في headers، فقط للتتبع الداخلي)

FONT_NAME = "Arial"
INK = "3B2A22"
NAVY_DARK = "5E4438"

def thin_border():
    s = Side(style="thin", color="D9CFC2")
    return Border(top=s, bottom=s, left=s, right=s)

def build_entries(reports, wastes):
    entries = []
    for rep in reports:
        ref = f"shift-report-{rep['id']}"
        detail = (
            f"تقرير وردية آلي — مبيعات {rep['total_sales']:,.0f} ج.م "
            f"({rep['invoices_count']} فاتورة)"
            + (f"، هدر {rep['waste_percentage']}%" if rep['waste_percentage'] is not None else "")
            + (f"، التزام ريسبي {rep['recipe_compliance_percentage']}%" if rep['recipe_compliance_percentage'] is not None else "")
            + (f"، سلامة غذاء {rep['food_safety_score']}%" if rep['food_safety_score'] is not None else "")
        )
        entries.append({
            "ref": ref,
            "sort_key": rep["created_at"] or "",
            "date": rep["report_date"],
            "time": rep["created_at"][11:16] if rep["created_at"] else "",
            "branch": rep["branch"],
            "source": "تطبيق Nest التفاعلي" if rep["source"] == "Nest API" else rep["source"],
            "type": "مبيعات-إيرادات",
            "detail": detail,
            "notes": rep["notes"] or "",
            "by": rep["manager_name"],
        })
    for w in wastes:
        ref = f"waste-entry-{w['id']}"
        detail = (
            f"تسجيل هدر آلي — {w['item_name']}: {w['quantity']:g} {w['unit']} "
            f"(السبب: {w['reason']})"
            + (f"، تكلفة تقديرية {w['estimated_cost']:,.0f} ج.م" if w['estimated_cost'] is not None else "")
        )
        entries.append({
            "ref": ref,
            "sort_key": w["created_at"] or "",
            "date": w["entry_date"],
            "time": w["created_at"][11:16] if w["created_at"] else "",
            "branch": w["branch"],
            "source": "تطبيق Nest التفاعلي" if w["source"] == "Nest API" else w["source"],
            "type": "مخزون",
            "detail": detail,
            "notes": w["notes"] or "",
            "by": w["recorded_by"],
        })
    entries.sort(key=lambda e: e["sort_key"])
    return entries


def main():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    reports = conn.execute("SELECT * FROM shift_reports ORDER BY id ASC").fetchall()
    wastes = conn.execute("SELECT * FROM waste_entries ORDER BY id ASC").fetchall()
    conn.close()

    entries = build_entries(reports, wastes)
    if not entries:
        print("لا توجد بيانات (تقارير وردية أو تسجيلات هدر) في قاعدة البيانات بعد.")
        return

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[LOG_SHEET]

    # اجمع معرّفات السجلات المُزامنة سابقاً (لتفادي التكرار)
    synced_ids = set()
    for row in ws.iter_rows(min_row=3, max_col=REF_COL):
        ref_cell = row[REF_COL - 1]
        if ref_cell.value:
            synced_ids.add(str(ref_cell.value))

    # ابحث عن أول صف فارغ فعلي (بعد آخر بيانات) — العمود B (التاريخ) هو المرجع
    next_row = 3
    for r in range(3, 501):
        if ws.cell(row=r, column=2).value not in (None, ""):
            next_row = r + 1

    added = 0
    for e in entries:
        if e["ref"] in synced_ids:
            continue
        r = next_row
        values = [
            None,  # # (يُملأ أدناه)
            e["date"],
            e["time"],
            e["branch"],
            e["source"],
            e["type"],
            e["detail"],
            e["notes"],
            e["by"],
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = Font(name=FONT_NAME, size=10, color=INK)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, readingOrder=2)
            cell.border = thin_border()
        ws.cell(row=r, column=1, value=r - 2)  # ترقيم متسق مع باقي الصفوف
        ws.cell(row=r, column=REF_COL, value=e["ref"])  # عمود التتبع المخفي
        next_row += 1
        added += 1

    if added:
        ws.column_dimensions["Z"].hidden = True
        wb.save(XLSX_PATH)
        print(f"تمت إضافة {added} صف/صفوف جديدة (تقارير وردية + تسجيلات هدر) إلى شيت السجل اليومي.")
    else:
        print("كل البيانات مُزامنة بالفعل — لا صفوف جديدة.")

if __name__ == "__main__":
    main()
